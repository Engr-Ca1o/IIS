import sys
import pandas as pd
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout,
                             QTableWidget, QTableWidgetItem, QPushButton, QLabel, QComboBox, 
                             QMessageBox, QHeaderView)
from PyQt5.QtGui import QFont, QPixmap, QIcon
from PyQt5.QtCore import Qt
from db_connector import connect_to_database

class AdminDashboard(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Admin Dashboard")
        self.setGeometry(50, 50, 800, 600)
        self.setWindowIcon(QIcon("mmd-logo.ico"))
        
        # Set a common font for all widgets
        self.font = QFont("Arial", 11)
        
        # Apply global styles
        self.setStyleSheet(
            "QWidget { background-color: #f0f8ff; }"
            "QLabel { font-size: 14pt; font-weight: bold; color: #0033cc; background-color: transparent; }"
            "QComboBox { border: 2px solid #0033cc; border-radius: 5px; padding: 5px; }"
            "QPushButton { background-color: #ffcc00; color: #0033cc; font-weight: bold; border: 2px solid #0033cc; border-radius: 5px; padding: 8px; }"
            "QPushButton:hover { background-color: #ffdb4d; }"
            "QTableWidget { border: 2px solid #0033cc; background-color: #ffffff; }"
        )

        main_layout = QVBoxLayout()

        # --------- Header Container with Gradient (only background) ---------
        header_container = QWidget()
        header_container.setStyleSheet(
            "background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1, stop: 0 #0814ce, stop: 1 #f0f8ff);"
        )
        header_layout = QHBoxLayout(header_container)
        
        # Logo label
        logo_label = QLabel()
        pixmap = QPixmap("mmd-logo.png")  # Replace with your logo file path
        logo_label.setPixmap(pixmap)
        logo_label.setScaledContents(True)
        logo_label.setFixedSize(100, 100)
        logo_label.setStyleSheet("background-color: transparent;")
        header_layout.addWidget(logo_label, alignment=Qt.AlignRight)
        
        # Title layout (vertical) for college name and departments
        title_layout = QVBoxLayout()
        title_label = QLabel("Holy Cross College")
        title_font = QFont("Old English Text MT", 20)
        title_label.setFont(title_font)
        title_label.setStyleSheet("font-weight: bold; font-size: 40pt; color: #ffffff; background-color: transparent;")
        title_label.setAlignment(Qt.AlignCenter)
        dept_label1 = QLabel("Information Communication Department")
        dept_label1.setStyleSheet("font-size: 12pt; color: #ffffff; background-color: transparent;")
        dept_label1.setAlignment(Qt.AlignCenter)
        dept_label2 = QLabel("Multimedia Department")
        dept_label2.setStyleSheet("font-size: 12pt; color: #ffffff; background-color: transparent;")
        dept_label2.setAlignment(Qt.AlignCenter)
        title_layout.addWidget(title_label)
        title_layout.addWidget(dept_label1)
        title_layout.addWidget(dept_label2)
        header_layout.addLayout(title_layout, stretch=1)
        header_layout.setAlignment(Qt.AlignCenter)
        # ----------------------------------------------------------------------
        
        main_layout.addWidget(header_container)
        # ----------------------------------------------------------------------

        # Filter Layout (remains without gradient)
        filter_layout = QHBoxLayout()
        filter_layout.setAlignment(Qt.AlignCenter)

        # Year filter group
        year_layout = QHBoxLayout()
        year_label = QLabel("Year:")
        year_label.setFont(self.font)
        self.year_filter = QComboBox()
        self.year_filter.setFont(self.font)
        self.year_filter.addItem("All Years")
        self.year_filter.addItems(["1st Year", "2nd Year", "3rd Year", "4th Year"])
        self.year_filter.currentIndexChanged.connect(self.load_data)
        year_layout.addWidget(year_label)
        year_layout.addWidget(self.year_filter)
        
        # Course filter group
        course_layout = QHBoxLayout()
        course_label = QLabel("Course:")
        course_label.setFont(self.font)
        self.course_filter = QComboBox()
        self.course_filter.setFont(self.font)
        self.course_filter.addItem("All Courses")
        self.course_filter.addItems(["Bachelor of Science in Accountancy", "Bachelor of Science in Accounting Information System",
                                     "Bachelor of Science in Civil Engineering", "Bachelor of Science in Computer Engineering",
                                     "Bachelor of Science in Criminology", "Bachelor of Science in Hospitality Management",
                                     "Bachelor of Science in Tourism Management", "Bachelor of Science in Information Technology",
                                     "Bachelor of Science in Computer Science", "Bachelor of Science in Psychology",
                                     "Bachelor of Science in Business Administration - Major in Financial Management", "Bachelor of Science in Business Administration - Major in Marketing Management",
                                     "Bachelor of Science in Business Administration - Major in Human Resources Management", "Bachelor of Science in Business Administration - Major in Operations Management",
                                     "Bachelor of Elementary Education", "Bachelor of Secondary Education - Major in English",
                                     "Bachelor of Secondary Education - Major in Mathematics", "Bachelor of Secondary Education - Major in Science",
                                     "Bachelor of Secondary Education - Major in Filipino", "Associate in Computer Technology",])
        self.course_filter.currentIndexChanged.connect(self.load_data)
        course_layout.addWidget(course_label)
        course_layout.addWidget(self.course_filter)
        
        # Sort order filter group
        sort_layout = QHBoxLayout()
        sort_label = QLabel("Sort by:")
        sort_label.setFont(self.font)
        self.sort_order = QComboBox()
        self.sort_order.setFont(self.font)
        self.sort_order.addItems(["Date Ascending", "Date Descending"])
        self.sort_order.currentIndexChanged.connect(self.load_data)
        sort_layout.addWidget(sort_label)
        sort_layout.addWidget(self.sort_order)
        
        filter_layout.addLayout(year_layout)
        filter_layout.addLayout(course_layout)
        filter_layout.addLayout(sort_layout)
        
        main_layout.addLayout(filter_layout)

        # Table widget
        self.table = QTableWidget()
        self.table.setFont(self.font)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        main_layout.addWidget(self.table)

        # Export to Excel button
        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.setFont(self.font)
        self.export_btn.clicked.connect(self.export_to_excel)
        main_layout.addWidget(self.export_btn)

        self.setLayout(main_layout)
        self.load_data()

    def load_data(self):
        conn = connect_to_database("admin")
        if conn is None:
            QMessageBox.critical(self, "Error", "Failed to connect to database.")
            return

        cursor = conn.cursor()  
        query = "SELECT * FROM students WHERE 1"
        params = []

        year = self.year_filter.currentText()
        if year != "All Years":
            query += " AND year = %s"
            params.append(year)

        course = self.course_filter.currentText()
        if course != "All Courses":
            query += " AND course = %s"
            params.append(course)

        order = self.sort_order.currentText()
        if order == "Date Ascending":
            query += " ORDER BY datetime ASC"
        else:
            query += " ORDER BY datetime DESC"

        cursor.execute(query, tuple(params))
        results = cursor.fetchall()
        conn.close()

        headers = ["Student ID", "Surname", "First Name", "MI", "Address",
                   "Year", "Course", "Emergency Name", "Relation to Student",
                   "Emergency Contact", "DateTime"]
        
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(results))
        for row_idx, row in enumerate(results):
            for col_idx, header in enumerate(headers):
                key = header.lower().replace(" ", "_")
                item = QTableWidgetItem(str(row.get(key, "")))
                item.setTextAlignment(Qt.AlignCenter)
                item.setFont(self.font)
                self.table.setItem(row_idx, col_idx, item)

    def export_to_excel(self):
        conn = connect_to_database()
        if conn is None:
            QMessageBox.critical(self, "Error", "Database connection failed.")
            return

        cursor = conn.cursor()
        query = "SELECT * FROM students"
        cursor.execute(query)
        data = cursor.fetchall()
        conn.close()

        if not data:
            QMessageBox.information(self, "No Data", "There is no data to export.")
            return

        df = pd.DataFrame(data)
        if "id" in df.columns:
            df.drop("id", axis=1, inplace=True)

        file_path = "student_data_export.xlsx"
        try:
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name="Data")
                worksheet = writer.sheets["Data"]

                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                for col_idx, col in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.alignment = header_alignment

                    max_length = max(df[col].astype(str).map(len).max(), len(col))
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                            min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            QMessageBox.information(self, "Export Successful", f"Data exported to {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export data:\n{e}")


def open_admin_dashboard():
    app = QApplication(sys.argv)
    window = AdminDashboard()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    open_admin_dashboard()
